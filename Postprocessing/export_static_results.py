import pandas as pd
import pypsa
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
import plotly.graph_objects as go
from matplotlib.lines import Line2D
import re
import networkx as nx
import matplotlib.pyplot as plt

def plot_energy_balance_sankey_static(dispatch_clean, grid):

    def safe_sum(df, col):
        return df[col].sum() if col in df.columns else 0.0

    pv = safe_sum(dispatch_clean, "PV")
    wind = safe_sum(dispatch_clean, "Wind")
    grid_import = safe_sum(dispatch_clean, "Grid_import")
    dispatch = safe_sum(dispatch_clean, "Dispatch")
    shedding = safe_sum(dispatch_clean, "shedding")

    load = grid.loads_t.p.sum().sum() if hasattr(grid.loads_t.p, "sum") else 0.0
    grid_export = -safe_sum(dispatch_clean, "Grid_export")

    served_load = max(load - shedding, 0.0)

    labels = [
        "PV",
        "Wind",
        "Grid import",
        "Dispatch",
        "Energy supplied",
        "Served load",
        "Grid export",
    ]

    source = [
        0, 1, 2, 3,   # inputs -> energy supplied
        4, 4          # energy supplied -> served load / grid export
    ]

    target = [
        4, 4, 4, 4,
        5, 6
    ]

    value = [
        pv,
        wind,
        grid_import,
        dispatch,
        served_load,
        grid_export
    ]

    links_df = pd.DataFrame({
        "source": source,
        "target": target,
        "value": value
    })

    links_df = links_df[links_df["value"] > 1e-9]

    fig = go.Figure(data=[go.Sankey(
        arrangement="snap",
        node=dict(
            pad=25,
            thickness=20,
            line=dict(color="black", width=0.5),
            label=labels
        ),
        link=dict(
            source=links_df["source"],
            target=links_df["target"],
            value=links_df["value"]
        )
    )])

    fig.update_layout(
        title=dict(
            text="Energy balance",
            x=0.03,
            y=0.95
        ),
        font=dict(size=16),
        height=750,
        margin=dict(l=20, r=20, t=100, b=20)
    )

    sankey_df = pd.DataFrame({
        "Category": [
            "",
            "",
            "PV",
            "Wind",
            "Dispatch",
            "Grid import",
            "Total supply",
            "",
            "",
            "Total demand",
            "Served load",
            "Unserved load",
            "Grid export",
            "Balance check",
        ],
        "Value (MW)": [
            "",
            "Supply side (MW)",
            pv,
            wind,
            dispatch,
            grid_import,
            pv + wind + dispatch + grid_import,
            "",
            "Demand side (MW)",
            load,
            served_load,
            shedding,
            grid_export,
            served_load + grid_export,
        ]
    })

    return fig, sankey_df


def insert_fig_in_sheet(ws, img_path, cell="A1", max_width=None, max_height=None):
    """
    Inserta una imagen en Excel manteniendo su proporción original.
    La imagen se escala para caber dentro de max_width x max_height.
    """

    img = XLImage(img_path)

    original_width = img.width
    original_height = img.height

    if max_width is None and max_height is None:
        ws.add_image(img, cell)
        return

    scale_w = max_width / original_width if max_width is not None else float("inf")
    scale_h = max_height / original_height if max_height is not None else float("inf")

    scale = min(scale_w, scale_h)

    if scale != float("inf"):
        img.width = original_width * scale
        img.height = original_height * scale

    ws.add_image(img, cell)



def save_fig(fig, path, dpi=150):
    if fig is None:
        return False

    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    return True


def save_plotly_html(fig, path):
    if fig is None:
        return False

    try:
        fig.write_html(str(path))
        return True
    except Exception as e:
        print(f"Error guardando HTML Plotly en {path}: {e}")
        return False
    

def save_plotly_fig(fig, path, width=1200, height=750, scale=2):
    if fig is None:
        return False
    try:
        fig.write_image(str(path), width=width, height=height, scale=scale)
        return True
    except Exception:
        return False


def drawGrid(grid: pypsa.Network, pcc_bus_name: str | None = None):
    """
    Devuelve una figura con la topología de la red diferenciando:
    - buses normales
    - buses de storage
    - PCC

    Esta versión NO usa grid.plot(), para evitar errores de PyPSA
    con pandas StringDtype en algunos entornos.
    """

    # -----------------------------
    # 1. Crear layout
    # -----------------------------
    G = grid.graph()
    pos = nx.spring_layout(G, seed=42)

    for bus in grid.buses.index:
        if bus in pos:
            grid.buses.loc[bus, "x"] = float(pos[bus][0])
            grid.buses.loc[bus, "y"] = float(pos[bus][1])

    bus_names = list(grid.buses.index)

    # -----------------------------
    # 2. Clasificar buses
    # -----------------------------
    if not grid.stores.empty and "bus" in grid.stores.columns:
        storage_buses = set(grid.stores["bus"].dropna().astype(str))
    else:
        storage_buses = set()

    if pcc_bus_name is not None and pcc_bus_name in bus_names:
        pcc_buses = {pcc_bus_name}
    else:
        pcc_buses = {bus for bus in bus_names if "PCC" in str(bus).upper()}

    normal_buses = set(map(str, bus_names)) - storage_buses - set(map(str, pcc_buses))

    # -----------------------------
    # 3. Crear figura
    # -----------------------------
    fig, ax = plt.subplots(figsize=(11.5, 8))

    # -----------------------------
    # 4. Dibujar líneas
    # -----------------------------
    if not grid.lines.empty:
        for _, row in grid.lines.iterrows():
            bus0 = row["bus0"]
            bus1 = row["bus1"]

            if bus0 in pos and bus1 in pos:
                x0, y0 = pos[bus0]
                x1, y1 = pos[bus1]

                ax.plot(
                    [x0, x1],
                    [y0, y1],
                    color="gray",
                    linewidth=1.8,
                    zorder=1
                )

    # -----------------------------
    # 5. Dibujar links
    # -----------------------------
    if hasattr(grid, "links") and not grid.links.empty:
        for _, row in grid.links.iterrows():
            bus0 = row["bus0"]
            bus1 = row["bus1"]

            if bus0 in pos and bus1 in pos:
                x0, y0 = pos[bus0]
                x1, y1 = pos[bus1]

                ax.plot(
                    [x0, x1],
                    [y0, y1],
                    color="orange",
                    linewidth=2.2,
                    linestyle="--",
                    zorder=2
                )

    # -----------------------------
    # 6. Dibujar buses manualmente
    # -----------------------------
    for bus in grid.buses.index:
        x = float(grid.buses.loc[bus, "x"])
        y = float(grid.buses.loc[bus, "y"])

        if bus in pcc_buses:
            color = "#C00000"
            size = 220
        elif str(bus) in storage_buses:
            color = "#70AD47"
            size = 170
        else:
            color = "#4F81BD"
            size = 190

        ax.scatter(
            x,
            y,
            s=size,
            c=color,
            edgecolors="black",
            linewidths=0.8,
            zorder=5
        )

    # -----------------------------
    # 7. Etiquetas
    # -----------------------------
    def bus_label(bus_name: str) -> str:
        name_upper = str(bus_name).upper()

        if "PCC" in name_upper:
            return "PCC"

        if str(bus_name) in storage_buses:
            matches = re.findall(r"\d+", str(bus_name))
            return f"S{matches[-1]}" if matches else str(bus_name)

        matches = re.findall(r"\d+", str(bus_name))
        return matches[-1] if matches else str(bus_name)

    for bus in grid.buses.index:
        x = float(grid.buses.loc[bus, "x"])
        y = float(grid.buses.loc[bus, "y"])

        if bus in pcc_buses:
            dx_text, dy_text = -0.02, -0.12
        elif str(bus) in storage_buses:
            dx_text, dy_text = 0.0, -0.11
        else:
            dx_text, dy_text = 0.0, -0.10

        ax.text(
            x + dx_text,
            y + dy_text,
            bus_label(bus),
            fontsize=11,
            ha="center",
            va="center",
            bbox=dict(
                facecolor="white",
                edgecolor="none",
                alpha=0.85,
                pad=1.4
            ),
            zorder=10
        )

    # -----------------------------
    # 8. Leyenda
    # -----------------------------
    legend_elements = [
        Line2D([0], [0], marker='o', color='w', label='Normal bus',
               markerfacecolor="#4F81BD", markeredgecolor="black", markersize=10),
        Line2D([0], [0], marker='o', color='w', label='Storage bus',
               markerfacecolor="#70AD47", markeredgecolor="black", markersize=10),
        Line2D([0], [0], marker='o', color='w', label='PCC bus',
               markerfacecolor="#C00000", markeredgecolor="black", markersize=10),
        Line2D([0], [0], color='gray', lw=2, label='Line'),
        Line2D([0], [0], color='orange', lw=2, linestyle='--', label='Link'),
    ]

    ax.legend(
        handles=legend_elements,
        loc="upper left",
        bbox_to_anchor=(1.02, 1.0),
        borderaxespad=0.0
    )

    ax.set_title("Grid topology")

    # -----------------------------
    # 9. Límites
    # -----------------------------
    x_vals = grid.buses["x"].astype(float)
    y_vals = grid.buses["y"].astype(float)

    x_min, x_max = x_vals.min(), x_vals.max()
    y_min, y_max = y_vals.min(), y_vals.max()

    dx = x_max - x_min
    dy = y_max - y_min

    if dx == 0:
        dx = 1.0
    if dy == 0:
        dy = 1.0

    ax.set_xlim(x_min - 0.22 * dx, x_max + 0.22 * dx)
    ax.set_ylim(y_min - 0.22 * dy, y_max + 0.14 * dy)

    ax.set_aspect("equal")
    ax.axis("off")

    plt.tight_layout()
    return fig


def build_dispatch_clean_static(grid: pypsa.Network) -> pd.DataFrame:
    dispatch_raw = grid.generators_t.p.copy()

    cols_base = dispatch_raw.columns.str.replace(r"_seg\d+$", "", regex=True)
    dispatch = dispatch_raw.T.groupby(cols_base).sum().T

    if any("PV" in c for c in dispatch.columns):
        dispatch["PV"] = dispatch[[c for c in dispatch.columns if "PV" in c]].sum(axis=1)

    if any("Wind" in c for c in dispatch.columns):
        dispatch["Wind"] = dispatch[[c for c in dispatch.columns if "Wind" in c]].sum(axis=1)

    if any("Dispatch" in c for c in dispatch.columns):
        dispatch["Dispatch"] = dispatch[[c for c in dispatch.columns if "Dispatch" in c]].sum(axis=1)

    if any("shedding" in c for c in dispatch.columns):
        dispatch["shedding"] = dispatch[[c for c in dispatch.columns if "shedding" in c]].sum(axis=1)

    # Export positivo en resultados
    if "Grid_export" in dispatch.columns:
        dispatch["Grid_export"] = -dispatch["Grid_export"]

    dispatch_clean = pd.DataFrame(index=dispatch.index)

    for col in ["PV", "Wind", "Dispatch", "Grid_import", "Grid_export", "shedding"]:
        if col in dispatch.columns:
            dispatch_clean[col] = dispatch[col]

    dispatch_clean = dispatch_clean.loc[:, (dispatch_clean.abs() > 1e-9).any()]

    return dispatch_clean


def export_static_results(grid: pypsa.Network, output_file: str = "results_static.xlsx") -> None:
    """
    Exporta a Excel los resultados relevantes de un OPF DC estático.

    Hojas generadas:
    - KPIs
    - dispatch
    - loads
    - line_flows
    - prices
    """

    def apply_borders(ws) -> None:
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in [None, ""]:
                    cell.border = border

    def autofit_columns(ws) -> None:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 35)

    # =========================================================
    # 1) DISPATCH DE GENERADORES
    # =========================================================
    dispatch_clean = build_dispatch_clean_static(grid)

    # =========================================================
    # 2) DEMANDA
    # =========================================================
    loads_df = grid.loads_t.p.copy()

    # =========================================================
    # 3) FLUJOS EN LÍNEAS
    # =========================================================
    line_flows = grid.lines_t.p0.copy()

    line_loading = pd.DataFrame(index=line_flows.index)
    for line in grid.lines.index:
        s_nom = grid.lines.loc[line, "s_nom"]
        if pd.notna(s_nom) and s_nom != 0:
            line_loading[line] = abs(line_flows[line]) / s_nom * 100.0
        else:
            line_loading[line] = 0.0

    # =========================================================
    # 4) PRECIOS NODALES
    # =========================================================
    prices = grid.buses_t.marginal_price.copy()

    # =========================================================
    # 5) KPIs / SANKEY
    # =========================================================
    img_dir = Path("results_static_figures")
    img_dir.mkdir(exist_ok=True)

    fig_sankey, df_sankey = plot_energy_balance_sankey_static(dispatch_clean, grid)
    fig_topology = drawGrid(grid)

    # =========================================================
    # 6) EXPORTAR A EXCEL
    # =========================================================
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_sankey.to_excel(writer, sheet_name="KPIs", index=False, header=False)

        dispatch_clean.round(4).to_excel(writer, sheet_name="dispatch")
        loads_df.round(4).to_excel(writer, sheet_name="loads")
        line_flows.round(4).to_excel(writer, sheet_name="line_flows")
        prices.round(4).to_excel(writer, sheet_name="prices")

        startrow = len(line_flows) + 3
        line_loading.round(4).to_excel(
            writer,
            sheet_name="line_flows",
            startrow=startrow
        )
    saved_sankey_html = save_plotly_html(fig_sankey, img_dir / "sankey.html")
    #saved_sankey = save_plotly_fig(fig_sankey, img_dir / "sankey.png")
    saved_topology = save_fig(fig_topology, img_dir / "gridtopology.png")
    # =========================================================
    # 7) FORMATO BÁSICO
    # =========================================================
    wb = load_workbook(output_file)

    for sheet in wb.sheetnames:
        autofit_columns(wb[sheet])

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.sheet_view.showGridLines = False
        apply_borders(ws)

    if "line_flows" in wb.sheetnames:
        ws_lines = wb["line_flows"]
        ws_lines.cell(row=len(line_flows) + 3, column=1).value = "Line loading (%)"
    """
    if saved_sankey:
        insert_fig_in_sheet(wb["KPIs"], img_dir / "sankey.png", cell="D2", max_width=380*2,
            max_height=260*2)
    """
    if saved_sankey_html:
        ws = wb["KPIs"]
        ws["B17"] = "Open interactive Sankey"
        ws["B17"].hyperlink = str((img_dir / "sankey.html").resolve())
        ws["B17"].style = "Hyperlink"
    
    if saved_topology:
        insert_fig_in_sheet(wb["KPIs"], img_dir / "gridtopology.png", cell="Q2", max_width=420*2,
            max_height=320*2)

    wb.save(output_file)
    wb.close()
