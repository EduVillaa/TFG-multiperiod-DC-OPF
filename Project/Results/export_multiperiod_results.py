import pandas as pd
import pypsa
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side
import kaleido
from matplotlib.lines import Line2D
import re
import networkx as nx
from Results.Graphs.dispatchgraphs import dispatch_graph_resolution_choice
from Results.Graphs.SOCgraphs import SOC_graph_resolution_choice
#from Results.Graphs.lineflowgraphs import maxloading_graph_resolution_choice, plot_line_loading_percent
from Results.Graphs.lineflowgraphs import *
from Results.Graphs.sankeygraph import plot_energy_balance_sankey
from Results.Graphs.renewablegraphs import renewable_graph_resolution_choice
from Results.Graphs.renewablesharegraphs import renewableshare_graph_resolution_choice
from Results.Graphs.import_export_graphs import GridExportImport_graph_resolution_choice
from Results.KPIsoptimized_battery import get_battery_sizes
from Results.Graphs.loadgraphs import total_load_graph_resolution_choice

def save_plotly_fig(fig, path, width=1200, height=750, scale=2):
    if fig is None:
        return False

    fig.write_image(str(path), width=width, height=height, scale=scale)
    return True

def save_fig(fig, path, dpi=150):
    if fig is None:
        return False

    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    return True


def insert_fig_in_sheet(ws, img_path, cell="A1", max_width=None, max_height=None):
    """
    Inserta una imagen en Excel manteniendo su proporción original.
    La imagen se escala para caber dentro de max_width x max_height.
    """

    img = XLImage(img_path)

    original_width = img.width
    original_height = img.height

    if max_width is None and max_height is None:
        ws.add_image(img, cell)
        return

    scale_w = max_width / original_width if max_width is not None else float("inf")
    scale_h = max_height / original_height if max_height is not None else float("inf")

    scale = min(scale_w, scale_h)

    if scale != float("inf"):
        img.width = original_width * scale
        img.height = original_height * scale

    ws.add_image(img, cell)


def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # letra de la columna

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length  # margen extra
        ws.column_dimensions[col_letter].width = adjusted_width

def apply_borders(ws):
    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                cell.border = border

def save_plotly_html(fig, path):
    if fig is None:
        return False

    fig.write_html(str(path))
    return True


def build_renewable_detailed_df(
    df_available_renewable: pd.DataFrame,
    gen_p_renewables: pd.DataFrame
) -> pd.DataFrame:
    """
    Devuelve un DataFrame con, para cada generador renovable:
    - available
    - used
    - curtailment

    y además tres columnas finales:
    - Total_available
    - Total_used
    - Total_curtailment
    """

    # Comprobaciones básicas
    if not df_available_renewable.index.equals(gen_p_renewables.index):
        raise ValueError("Los índices temporales de ambos DataFrames no coinciden.")

    if list(df_available_renewable.columns) != list(gen_p_renewables.columns):
        raise ValueError("Las columnas de ambos DataFrames no coinciden ni en nombre ni en orden.")

    df_result = pd.DataFrame(index=df_available_renewable.index)

    for col in df_available_renewable.columns:
        available_col = df_available_renewable[col]
        used_col = gen_p_renewables[col]
        curtailment_col = available_col - used_col

        df_result[f"{col}_available"] = available_col
        df_result[f"{col}_used"] = used_col
        df_result[f"{col}_curtailment"] = curtailment_col

    # Totales
    df_result["Total_available"] = df_available_renewable.sum(axis=1)
    df_result["Total_used"] = gen_p_renewables.sum(axis=1)
    df_result["Total_curtailment"] = (
        df_result["Total_available"] - df_result["Total_used"]
    )

    return df_result



def drawGrid(grid: pypsa.Network, pcc_bus_name: str | None = None):
    """
    Devuelve una figura con la topología de la red diferenciando:
    - buses normales
    - buses de storage
    - PCC
    """

    # -----------------------------
    # 1. Crear layout
    # -----------------------------
    G = grid.graph()
    pos = nx.spring_layout(G, seed=42)
    # pos = nx.kamada_kawai_layout(G)
    # pos = nx.circular_layout(G)

    for bus in grid.buses.index:
        if bus in pos:
            grid.buses.loc[bus, "x"] = pos[bus][0]
            grid.buses.loc[bus, "y"] = pos[bus][1]

    bus_names = list(grid.buses.index)

    # -----------------------------
    # 2. Clasificar buses
    # -----------------------------
    if not grid.stores.empty and "bus" in grid.stores.columns:
        storage_buses = set(grid.stores["bus"].dropna())
    else:
        storage_buses = set()

    if pcc_bus_name is not None and pcc_bus_name in bus_names:
        pcc_buses = {pcc_bus_name}
    else:
        pcc_buses = {bus for bus in bus_names if "PCC" in str(bus).upper()}

    normal_buses = set(bus_names) - storage_buses - pcc_buses

    # -----------------------------
    # 3. Colores y tamaños por bus
    # -----------------------------
    bus_colors = []
    bus_sizes = []

    for bus in grid.buses.index:
        if bus in pcc_buses:
            bus_colors.append("#C00000")   # rojo
            bus_sizes.append(0.010)        # más grande
        elif bus in storage_buses:
            bus_colors.append("#70AD47")   # verde
            bus_sizes.append(0.007)        # un poco más pequeño
        else:
            bus_colors.append("#4F81BD")   # azul
            bus_sizes.append(0.0085)       # normal

    # -----------------------------
    # 4. Dibujar red
    # -----------------------------
    fig, ax = plt.subplots(figsize=(11.5, 8))

    grid.plot(
        ax=ax,
        bus_sizes=bus_sizes,
        bus_colors=bus_colors,
        line_colors="gray",
        link_colors="orange"
    )

    # -----------------------------
    # 5. Etiquetas
    # -----------------------------
    def bus_label(bus_name: str) -> str:
        name_upper = str(bus_name).upper()

        if "PCC" in name_upper:
            return "PCC"

        if bus_name in storage_buses:
            matches = re.findall(r"\d+", str(bus_name))
            return f"S{matches[-1]}" if matches else str(bus_name)

        matches = re.findall(r"\d+", str(bus_name))
        return matches[-1] if matches else str(bus_name)

    for bus in grid.buses.index:
        x = float(grid.buses.loc[bus, "x"])
        y = float(grid.buses.loc[bus, "y"])

        # offsets distintos según tipo de bus
        if bus in pcc_buses:
            dx_text, dy_text = -0.02, -0.12
        elif bus in storage_buses:
            dx_text, dy_text = 0.0, -0.11
        else:
            dx_text, dy_text = 0.0, -0.10

        ax.text(
            x + dx_text,
            y + dy_text,
            bus_label(bus),
            fontsize=11,
            ha="center",
            va="center",
            bbox=dict(
                facecolor="white",
                edgecolor="none",
                alpha=0.85,
                pad=1.4
            ),
            zorder=10
        )

    # -----------------------------
    # 6. Leyenda fuera del gráfico
    # -----------------------------
    legend_elements = [
        Line2D([0], [0], marker='o', color='w', label='Normal bus',
               markerfacecolor="#4F81BD", markersize=10),
        Line2D([0], [0], marker='o', color='w', label='Storage bus',
               markerfacecolor="#70AD47", markersize=10),
        Line2D([0], [0], marker='o', color='w', label='PCC bus',
               markerfacecolor="#C00000", markersize=10),
        Line2D([0], [0], color='gray', lw=2, label='Line'),
        Line2D([0], [0], color='orange', lw=2, label='Link'),
    ]

    ax.legend(
        handles=legend_elements,
        loc="upper left",
        bbox_to_anchor=(1.02, 1.0),
        borderaxespad=0.0
    )

    ax.set_title("Grid topology")

    # -----------------------------
    # 7. Forzar límites para evitar cortes
    # -----------------------------
    x_vals = grid.buses["x"].astype(float)
    y_vals = grid.buses["y"].astype(float)

    x_min, x_max = x_vals.min(), x_vals.max()
    y_min, y_max = y_vals.min(), y_vals.max()

    dx = x_max - x_min
    dy = y_max - y_min

    if dx == 0:
        dx = 1.0
    if dy == 0:
        dy = 1.0

    ax.set_xlim(x_min - 0.22 * dx, x_max + 0.22 * dx)
    ax.set_ylim(y_min - 0.22 * dy, y_max + 0.14 * dy)

    plt.tight_layout()

    return fig

def export_multiperiod_results(grid: pypsa.Network, df_SYS_settings: pd.DataFrame, df_available_renewable: pd.DataFrame) -> None:
    # Agrupar generadores PWL
    cols_base = grid.generators_t.p.columns.str.replace(r'_seg\d+$', '', regex=True)
    dispatch = grid.generators_t.p.T.groupby(cols_base).sum().T
    dispatch["PV"] = dispatch[[c for c in dispatch.columns if "PV" in c]].sum(axis=1)
    dispatch["Wind"] = dispatch[[c for c in dispatch.columns if "Wind" in c]].sum(axis=1)
    dispatch["Dispatch"] = dispatch[[c for c in dispatch.columns if "Dispatch" in c]].sum(axis=1)
    dispatch["shedding"] = dispatch[[c for c in dispatch.columns if "shedding" in c]].sum(axis=1)

    charge_cols = [c for c in grid.links_t.p0.columns if c.startswith("BatteryCharge_")]
    discharge_cols = [c for c in grid.links_t.p1.columns if c.startswith("BatteryDischarge_")]
    
    # lado AC
    battery_charge = grid.links_t.p0[charge_cols].clip(lower=0).sum(axis=1)
    battery_discharge = (-grid.links_t.p1[discharge_cols]).clip(lower=0).sum(axis=1)

    dispatch.insert(0, "battery_discharge", battery_discharge)
    dispatch.insert(1, "battery_charge", -battery_charge)

    if "Grid_export" in dispatch.columns:
        dispatch["Grid_export"] = -dispatch["Grid_export"]

    dispatch_clean = pd.DataFrame(index=dispatch.index)
    for col in [
        "PV",
        "Wind",
        "battery_discharge",
        "Dispatch",
        "Grid_import",
        "battery_charge",
        "Grid_export",
        "shedding",
    ]:
        if col in dispatch.columns:
            dispatch_clean[col] = dispatch[col]

    dispatch_clean = dispatch_clean.loc[:, (dispatch_clean.abs() > 1e-6).any()]

    # Carpeta de imágenes
    img_dir = Path("results_figures")
    img_dir.mkdir(exist_ok=True)

    # Generar figuras
    fig_dispatch = dispatch_graph_resolution_choice(df_SYS_settings, dispatch_clean)
    fig_soc_total, fig_soc_batteries = SOC_graph_resolution_choice(df_SYS_settings, grid)
    fig_line_heatmap, fig_line_loading = maxloading_graph_resolution_choice(df_SYS_settings, grid)
    fig_max_line_loading = plot_line_loading_percent(grid, "Multiperiod")
    fig_total_loading_histogram = plot_line_loading_histogram_global(grid, "Multiperiod")
    fig_top_lines_loading_histogram = plot_line_loading_histogram_top_lines(grid, "Multiperiod", 3)
    fig_sankey, df_sankey = plot_energy_balance_sankey(dispatch_clean, grid, df_available_renewable)
    fig_total_load = total_load_graph_resolution_choice(df_SYS_settings, grid)
    fig_export_import = GridExportImport_graph_resolution_choice(df_SYS_settings, dispatch_clean)
    fig_renewable_total, fig_renewable_pv_wind = renewable_graph_resolution_choice(df_SYS_settings, 
                                                                                   dispatch_clean, df_available_renewable)
    fig_renewable_share_total = renewableshare_graph_resolution_choice(df_SYS_settings, dispatch_clean, grid)
    fig_grid_topology = drawGrid(grid)

    df_bat_optimized_data = get_battery_sizes(grid)

    gen_p_renewables = grid.generators_t.p.loc[:, 
    grid.generators_t.p.columns.str.contains("PV|Wind")
    ]
    print("df_available_renewable columns:")
    print(df_available_renewable.columns.tolist())

    print("gen_p_renewables columns:")
    print(gen_p_renewables.columns.tolist())

    print("Only in df_available_renewable:",
    set(df_available_renewable.columns) - set(gen_p_renewables.columns))

    print("Only in gen_p_renewables:",
    set(gen_p_renewables.columns) - set(df_available_renewable.columns))

    df_detailed_renewable = build_renewable_detailed_df(
    df_available_renewable,
    gen_p_renewables
    )   

    df_loads = pd.concat(
    [grid.loads_t.p, grid.loads_t.p.sum(axis=1).rename("Total_load")],
    axis=1
)

    # Guardar Excel con tablas
    output_file = "results_multiperiod.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_sankey.to_excel(writer, sheet_name="KPIs", index=False, header=False, startcol=1)

        df_bat_optimized_data.to_excel(
        writer,
        sheet_name="KPIs",
        startrow=1,   
        index=True,
        startcol=5,
    )
        dispatch_clean.round(2).to_excel(writer, sheet_name="dispatch")
        df_detailed_renewable.round(2).to_excel(writer, sheet_name="renewables")
        grid.stores_t.e.round(2).to_excel(writer, sheet_name="battery soc")
        grid.lines_t.p0.round(2).to_excel(writer, sheet_name="line flows")
        grid.buses_t.marginal_price.round(2).to_excel(writer, sheet_name="prices")
        df_loads.to_excel(writer, sheet_name="loads")

    # Guardar PNGs
    saved_dispatch = save_fig(fig_dispatch, img_dir / "dispatch.png")
    saved_sankey = save_plotly_fig(fig_sankey, img_dir / "sankey.png")
    saved_sankey_html = save_plotly_html(fig_sankey, img_dir / "sankey.html")
    saved_soc_total = save_fig(fig_soc_total, img_dir / "battery_soc_total.png")
    saved_soc_batteries = save_fig(fig_soc_batteries, img_dir / "battery_soc_batteries.png")

    saved_line_loading = save_fig(fig_line_loading, img_dir / "line_loading.png")
    saved_max_line_loading = save_fig(fig_max_line_loading, img_dir / "max_line_loading.png")
    saved_line_heatmap = save_fig(fig_line_heatmap, img_dir / "line_loading_heatmap.png")
    saved_total_loading_histogram = save_fig(fig_total_loading_histogram, img_dir / "total_loading_histogram.png")
    saved_top_loaded_lines_histogram = save_fig(fig_top_lines_loading_histogram, img_dir / "top_loaded_lines_histogram.png")

    saved_export_import = save_fig(fig_export_import, img_dir / "export_import.png")
    saved_total_load = save_fig(fig_total_load, img_dir / "total_load.png")

    saved_renewable_total = save_fig(fig_renewable_total, img_dir / "renewable_total.png")
    saved_renewable_pv_wind = save_fig(fig_renewable_pv_wind, img_dir / "renewable.png")

    saved_renewable_share_total = save_fig(fig_renewable_share_total, img_dir / "renewable_share.png")

    saved_grid_topology = save_fig(fig_grid_topology, img_dir / "gridtopology.png")

    # Reabrir workbook e insertar imágenes
    wb = load_workbook(output_file)

    for sheet in wb.sheetnames:
        autofit_columns(wb[sheet])

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.sheet_view.showGridLines = False
        apply_borders(ws)


    if saved_total_load:
        insert_fig_in_sheet(
            wb["loads"],
            img_dir / "total_load.png",
            cell="A30",
            max_width=520*2,
            max_height=300*2
        )

    if saved_renewable_total:
        insert_fig_in_sheet(
            wb["renewables"],
            img_dir / "renewable_total.png",
            cell="A30",
            max_width=520*2,
            max_height=300*2
        )

    if saved_renewable_pv_wind:
        insert_fig_in_sheet(
            wb["renewables"],
            img_dir / "renewable.png",
            cell="J30",
            max_width=520*2,
            max_height=300*2
        )

    if saved_renewable_share_total:
        insert_fig_in_sheet(
            wb["renewables"],
            img_dir / "renewable_share.png",
            cell="J6",
            max_width=520*2,
            max_height=300*2
        )

    if saved_dispatch:
        insert_fig_in_sheet(
            wb["dispatch"],
            img_dir / "dispatch.png",
            cell="K2",
            max_width=520*2,
            max_height=300*2
        )

    if saved_sankey:
        insert_fig_in_sheet(
            wb["KPIs"],
            img_dir / "sankey.png",
            cell="F10",
            max_width=380*2,
            max_height=260*2
        )

    if saved_export_import:
        insert_fig_in_sheet(
            wb["dispatch"],
            img_dir / "export_import.png",
            cell="K26",
            max_width=520*2,
            max_height=300*2
        )

    if saved_soc_total:
        insert_fig_in_sheet(
            wb["battery soc"],
            img_dir / "battery_soc_total.png",
            cell="J2",
            max_width=520*2,
            max_height=300*2
        )

    if saved_soc_batteries:
        insert_fig_in_sheet(
            wb["battery soc"],
            img_dir / "battery_soc_batteries.png",
            cell="J25",
            max_width=520*2,
            max_height=300*2
        )

    if saved_line_loading:
        insert_fig_in_sheet(
            wb["line flows"],
            img_dir / "line_loading.png",
            cell="A32",
            max_width=520*2,
            max_height=300*2
        )

    if saved_max_line_loading:
        insert_fig_in_sheet(
            wb["line flows"],
            img_dir / "max_line_loading.png",
            cell="Q45",
            max_width=520*2,
            max_height=300*2
        )

    if saved_line_heatmap:
        insert_fig_in_sheet(
            wb["line flows"],
            img_dir / "line_loading_heatmap.png",
            cell="A13",
            max_width=520*2,
            max_height=320*2
        )

    if saved_total_loading_histogram:
        insert_fig_in_sheet(
            wb["line flows"],
            img_dir / "total_loading_histogram.png",
            cell="Q1",
            max_width=420*2,
            max_height=260*2
        )

    if saved_top_loaded_lines_histogram:
        insert_fig_in_sheet(
            wb["line flows"],
            img_dir / "top_loaded_lines_histogram.png",
            cell="Q23",
            max_width=420*2,
            max_height=260*2
        )

    if saved_grid_topology:
        insert_fig_in_sheet(
            wb["KPIs"],
            img_dir / "gridtopology.png",
            cell="L10",
            max_width=420*2,
            max_height=320*2
        )

    wb.save(output_file)
    wb.close()
